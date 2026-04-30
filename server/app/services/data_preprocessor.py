"""
Data preprocessing pipeline — runs BEFORE DuckDB sampling and Parquet conversion.

Handles every supported file type:
    .csv  .txt  .tsv  .tab  .xlsx  .xls  .xlsm  .xlsb

Resource model (zero-disk for CSV, half-disk for Excel)
────────────────────────────────────────────────────────
File type   Disk used during preprocessing       RAM (peak)
─────────   ──────────────────────────────       ──────────
CSV / text  ZERO  — probe bytes via HTTP range   one chunk  (~200 MB)
            read, stream full file from Azure,   regardless of
            write clean blocks directly back      file size
            to Azure. No local file ever written.

Excel       ONE temp file (the .xlsx download). Openpyxl needs a seekable
            file. Rows are streamed via openpyxl read_only=True and written
            directly to Azure block blob — no second local copy of clean CSV.

Memory model
────────────
Large files (> 50 MB) are streamed 100 000 rows at a time. Small files
(<= 50 MB) are kept in RAM to allow exact-duplicate removal.

Cleaning stages (per-chunk for large files, once for small)
────────────────────────────────────────────────────────────
  1.  Encoding detection   — HTTP range read of first 64 KB only
  2.  Delimiter detection  — first 8 KB of probe bytes
  3.  Header detection     — first 15 rows of probe bytes
  4.  Schema discovery     — first 1 000 rows of probe bytes → per-column type
  5.  Per-chunk streaming:
        a. String cleaning  — BOM / control-chars / invisible-unicode / whitespace
        b. Null normalise   — "", "NULL", "N/A", "nan", ... -> ""
        c. Garbage-row drop — subtotals, separator lines, fully-empty rows
        d. Per-column conv  — bool / date / numeric normalisation
  6.  Duplicate removal    — ONLY for small files; skipped for large
  7.  Output              — written as Azure Block Blob (staged 4 MB blocks)

Output: clean UTF-8 CSV at preprocessed/{file_id}_clean.csv in same container.
"""
from __future__ import annotations

import asyncio
import base64
import csv
import io
import os
import re
import tempfile
import time
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timedelta, date as _date_type
from pathlib import Path
from typing import Callable, Iterator

import numpy as np
import pandas as pd
from azure.storage.blob import BlobClient, BlobServiceClient

from app.core.logger import ingest_logger


# ── Tunable constants ─────────────────────────────────────────────────────────

CHUNK_ROWS              = 100_000       # rows per streaming chunk
SMALL_FILE_THRESHOLD_MB = 50            # files under this get full-load + dedup
HEADER_SCAN_ROWS        = 15            # max rows to scan for the real header
TYPE_DETECT_SAMPLE_ROWS = 1_000         # rows used for column-type detection
DATE_CONVERT_THRESHOLD  = 0.55          # fraction that must parse as date
NUM_CONVERT_THRESHOLD   = 0.75          # fraction that must parse as numeric
NUM_HINT_THRESHOLD      = 0.50          # lower threshold when col name is a hint
PROBE_BYTES             = 256 * 1024    # bytes range-read from Azure for probing
BLOCK_SIZE              = 4 * 1024 * 1024  # 4 MB per Azure block blob block


# ── Supported file-type groups ────────────────────────────────────────────────

EXCEL_EXTS = frozenset({".xlsx", ".xls", ".xlsm", ".xlsb"})
TEXT_EXTS  = frozenset({".csv", ".txt", ".tsv", ".tab"})
ALL_EXTS   = EXCEL_EXTS | TEXT_EXTS


# ── Null-like string patterns (compared after .strip().lower()) ───────────────

_NULLSTR: frozenset[str] = frozenset({
    "", "null", "none", "na", "n/a", "nan", "nil", "tbd", "n.a.", "n.a",
    "-", "--", "---", ".", "..", "?", "#", "#n/a", "#na", "#null!",
    "not available", "not applicable", "not provided", "not assigned",
    "missing", "unknown", "no data", "no value", "nd", "n.d.",
    "void", "blank", "empty",
})


# ── Boolean lookup tables ─────────────────────────────────────────────────────

_BOOL_TRUE  = frozenset({"yes", "y", "true",  "t", "1", "on",  "enabled",  "active",   "x", "\u2713"})
_BOOL_FALSE = frozenset({"no",  "n", "false", "f", "0", "off", "disabled", "inactive", " ", "\u2717"})
_BOOL_ALL   = _BOOL_TRUE | _BOOL_FALSE


# ── Column-name heuristic word sets ──────────────────────────────────────────

_DATE_NAME_CLUES: frozenset[str] = frozenset({
    "date", "dt", "time", "timestamp", "created", "updated", "modified",
    "period", "dob", "birth", "expir", "effective", "since", "until",
    "_at", "at_", "start", "end", "from", "to", "year", "month", "week",
    "day", "posted", "issued", "received", "shipped", "closed",
})
_NUM_NAME_CLUES: frozenset[str] = frozenset({
    "amount", "price", "cost", "total", "sum", "count", "qty", "quantity",
    "balance", "rate", "pct", "percent", "ratio", "score", "revenue",
    "profit", "loss", "tax", "fee", "charge", "salary", "wage", "budget",
    "num", "number", "no.", "vol", "volume",
})


# ── Compiled regex patterns ───────────────────────────────────────────────────

_GARBAGE_ROW_RE = re.compile(
    r"^\s*(total|grand\s+total|subtotal|sub\s+total|sum|page\s+total|"
    r"running\s+total|end\s+of\s+report|average|avg|mean|balance\s+forward|"
    r"carried\s+forward|min|max)\b",
    re.IGNORECASE,
)
_SEP_ROW_RE    = re.compile(r"^[-=*_~\s|+]+$")
_CURRENCY_RE   = re.compile(r"[$\u20b9\u20ac\xa3\xa5\u20a9\u20a6\u20b1\u20ba\u20b4\u20bd\xa2\u0e3f]+")
_SPACE_THOU_RE = re.compile(r"(\d)\s(\d)")
_COMMA_THOU_RE = re.compile(r"^[+-]?[\d,]+\.?\d*$")
_PERCENT_RE    = re.compile(r"^([+-]?\d+\.?\d*)\s*%$")
_CTRL_RE       = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_INVISIBLE_RE  = re.compile(
    r"[\u200b\u200c\u200d\u200e\u200f\ufeff\u00ad"
    r"\u180e\u2060\u2061\u2062\u2063\u2064\u3000]"
)


# ── Excel date serial constants ───────────────────────────────────────────────

_EXCEL_EPOCH      = datetime(1899, 12, 30)
_EXCEL_SERIAL_MIN = 7300    # ~Jan 1920
_EXCEL_SERIAL_MAX = 73000   # ~Dec 2099


# ── Result dataclass ──────────────────────────────────────────────────────────

@dataclass
class PreprocessResult:
    clean_blob_path: str
    original_rows:   int
    clean_rows:      int
    rows_dropped:    int
    cols_renamed:    dict = field(default_factory=dict)
    warnings:        list = field(default_factory=list)
    encoding:        str  = "utf-8"
    file_type:       str  = "csv"
    used_streaming:  bool = False


# ══════════════════════════════════════════════════════════════════════════════
# Azure streaming helpers
# ══════════════════════════════════════════════════════════════════════════════

class _AzureRawStream(io.RawIOBase):
    """
    Wraps an Azure StorageStreamDownloader as a readable io.RawIOBase so
    pandas can read it directly via pd.read_csv(stream).

    No data is written to disk.  The Azure SDK fetches data over HTTP in
    chunks; this class stitches those chunks together into the read()
    interface that pandas expects.
    """

    def __init__(self, downloader) -> None:
        self._chunks   = downloader.chunks()
        self._leftover = b""

    def readable(self) -> bool:
        return True

    def readinto(self, b: bytearray) -> int:
        want = len(b)
        # Keep pulling Azure chunks until we have enough bytes or stream ends
        while len(self._leftover) < want:
            try:
                self._leftover += next(self._chunks)
            except StopIteration:
                break
        n = min(want, len(self._leftover))
        if n == 0:
            return 0
        b[:n] = self._leftover[:n]
        self._leftover = self._leftover[n:]
        return n


class _BlockBlobWriter:
    """
    Accumulates bytes in an in-memory buffer and uploads to Azure Blob Storage
    using the Block Blob API (stage_block + commit_block_list).

    Why Block Blob?
        • We can upload arbitrarily large outputs without holding the full
          content in memory or on disk.
        • Each block is at most BLOCK_SIZE bytes, so peak extra RAM is
          bounded by one block (~4 MB).
        • commit() finalises the blob atomically.

    Usage:
        writer = _BlockBlobWriter(blob_client)
        writer.write(b"some bytes")
        ...
        writer.commit()   # must be called exactly once at the end
    """

    def __init__(self, bc: BlobClient) -> None:
        self._bc     = bc
        self._buf    = bytearray()
        self._blocks: list[str] = []
        self._idx    = 0

    def write(self, data: bytes) -> None:
        self._buf.extend(data)
        while len(self._buf) >= BLOCK_SIZE:
            self._flush_block(bytes(self._buf[:BLOCK_SIZE]))
            del self._buf[:BLOCK_SIZE]

    def _flush_block(self, data: bytes) -> None:
        block_id = base64.b64encode(f"{self._idx:08d}".encode()).decode()
        self._bc.stage_block(block_id=block_id, data=data)
        self._blocks.append(block_id)
        self._idx += 1

    def commit(self) -> None:
        """Flush remaining buffer and commit all staged blocks."""
        if self._buf:
            self._flush_block(bytes(self._buf))
            self._buf = bytearray()
        if self._blocks:
            self._bc.commit_block_list(self._blocks)
        else:
            # Nothing was written — produce an empty blob with just the header
            self._bc.upload_blob(b"", overwrite=True)


# ══════════════════════════════════════════════════════════════════════════════
# Public entry point
# ══════════════════════════════════════════════════════════════════════════════

async def preprocess_file(
    blob_path:         str,
    file_name:         str,
    file_id:           str,
    connection_string: str,
    container_name:    str,
) -> PreprocessResult:
    """
    Clean a raw file and write clean CSV output back to Azure Blob Storage.

    Disk usage:
        CSV / text : ZERO disk.  Probe bytes via HTTP range-read (256 KB),
                     full file streamed directly from Azure, clean output
                     written as Azure Block Blob — no local file at any point.
        Excel      : ONE temp file (the .xlsx download; openpyxl needs
                     seekable I/O). Clean output still goes directly to Azure
                     as a Block Blob — no second local copy.

    RAM usage:
        Bounded by CHUNK_ROWS (100 000 rows ~ 200 MB) for any file size.
        For small files (<= 50 MB) the full dataframe is kept for dedup.

    Output blob: preprocessed/{file_id}_clean.csv
    """
    t0 = time.perf_counter()
    warns: list[str] = []
    ext       = Path(file_name).suffix.lower()
    file_type = "excel" if ext in EXCEL_EXTS else "csv"

    ingest_logger.info("preprocess", status="started", blob_path=blob_path,
                       file_name=file_name, file_type=file_type)

    # ── Probe size without downloading the file ───────────────────────────────
    src_bc    = await asyncio.to_thread(_get_blob_client, connection_string, container_name, blob_path)
    props     = await asyncio.to_thread(lambda: src_bc.get_blob_properties())
    file_size = props["size"]
    size_mb   = file_size / (1024 * 1024)
    is_large  = size_mb > SMALL_FILE_THRESHOLD_MB

    ingest_logger.info("preprocess", status="probed",
                       size_mb=round(size_mb, 1), streaming=is_large)

    clean_blob_path = f"preprocessed/{file_id}_clean.csv"
    dst_bc          = await asyncio.to_thread(
        _get_blob_client, connection_string, container_name, clean_blob_path
    )
    block_writer    = _BlockBlobWriter(dst_bc)

    if ext in EXCEL_EXTS:
        # Excel still needs one local temp file (openpyxl requires seekable I/O)
        with tempfile.TemporaryDirectory() as tmpdir:
            raw_path = os.path.join(tmpdir, f"raw{ext}")
            await asyncio.to_thread(
                _download_blob_to_file,
                blob_path, container_name, connection_string, raw_path,
            )
            result = await asyncio.to_thread(
                _process_excel_to_blob, raw_path, block_writer, ext, is_large, warns,
            )
        # tmpdir (and the only temp file) is deleted here; block_writer still in RAM
    else:
        # CSV/text: fully streaming, zero disk
        result = await asyncio.to_thread(
            _process_text_stream,
            src_bc, block_writer, ext, file_size, is_large, warns,
        )

    ingest_logger.info("preprocess", status="cleaned",
                       original_rows=result["original_rows"],
                       clean_rows=result["clean_rows"],
                       rows_dropped=result["original_rows"] - result["clean_rows"],
                       streaming=is_large)

    ingest_logger.info("preprocess", status="done",
                       clean_blob_path=clean_blob_path,
                       duration_ms=round((time.perf_counter() - t0) * 1000, 1))

    return PreprocessResult(
        clean_blob_path=clean_blob_path,
        original_rows=result["original_rows"],
        clean_rows=result["clean_rows"],
        rows_dropped=result["original_rows"] - result["clean_rows"],
        cols_renamed=result["cols_renamed"],
        warnings=warns,
        encoding=result.get("encoding", "utf-8"),
        file_type=file_type,
        used_streaming=is_large,
    )


# ══════════════════════════════════════════════════════════════════════════════
# Azure client helpers
# ══════════════════════════════════════════════════════════════════════════════

def _get_blob_client(conn_str: str, container: str, blob_path: str) -> BlobClient:
    return BlobServiceClient.from_connection_string(conn_str).get_blob_client(
        container=container, blob=blob_path
    )


def _probe_blob(src_bc: BlobClient, length: int = PROBE_BYTES) -> bytes:
    """Range-read the first `length` bytes without downloading the entire blob."""
    actual = min(length, src_bc.get_blob_properties()["size"])
    return src_bc.download_blob(offset=0, length=actual).readall()


def _download_blob_to_file(blob_path: str, container: str, conn_str: str, dest: str) -> None:
    """Full download to a local file (used only for Excel)."""
    bc = _get_blob_client(conn_str, container, blob_path)
    with open(dest, "wb") as fh:
        bc.download_blob().readinto(fh)


# ══════════════════════════════════════════════════════════════════════════════
# CSV / text — fully streaming, zero disk
# ══════════════════════════════════════════════════════════════════════════════

def _process_text_stream(
    src_bc:       BlobClient,
    block_writer: _BlockBlobWriter,
    ext:          str,
    file_size:    int,
    is_large:     bool,
    warns:        list[str],
) -> dict:
    """
    CSV/text processing: Azure -> RAM chunks -> Azure block blob.
    No local file is written at any point.

    Pass 1 (probe): HTTP range-read of first PROBE_BYTES (256 KB) →
        detect encoding, delimiter, header row, column types.
    Pass 2 (full): full Azure stream → pd.read_csv(chunksize) →
        clean each chunk → stage as block blob blocks.
    """
    # ── Pass 1: probe ─────────────────────────────────────────────────────────
    probe      = _probe_blob(src_bc, PROBE_BYTES)
    encoding   = _detect_encoding_from_bytes(probe)
    probe_text = probe.decode(encoding, errors="replace")
    delimiter  = _detect_delimiter_from_str(probe_text, ext)

    head_df = pd.read_csv(
        io.StringIO(probe_text), sep=delimiter, header=None, dtype=str,
        keep_default_na=False, nrows=HEADER_SCAN_ROWS, on_bad_lines="skip",
    )
    head_df = head_df.apply(_clean_str_series).apply(_nullify_series)
    header_row_idx = _find_header_row(head_df)

    raw_headers = [
        _flatten_col_name(v) or f"col_{i}"
        for i, v in enumerate(head_df.iloc[header_row_idx])
    ]
    headers, cols_renamed = _dedup_column_names(raw_headers)
    skip_rows = header_row_idx + 1

    if header_row_idx > 0:
        warns.append(
            f"Header row found at row {header_row_idx} ({header_row_idx} leading rows skipped)"
        )

    sample_df = pd.read_csv(
        io.StringIO(probe_text), sep=delimiter, header=None, dtype=str, names=headers,
        keep_default_na=False, skiprows=skip_rows,
        nrows=TYPE_DETECT_SAMPLE_ROWS, on_bad_lines="skip",
    )
    sample_df  = sample_df.apply(_clean_str_series).apply(_nullify_series)
    converters = _build_converters(sample_df, headers, warns)

    # ── Pass 2: full streaming read ────────────────────────────────────────────
    downloader = src_bc.download_blob()
    raw_stream = io.BufferedReader(_AzureRawStream(downloader), buffer_size=8 * 1024 * 1024)

    reader = pd.read_csv(
        raw_stream, sep=delimiter, header=None, dtype=str, names=headers,
        encoding=encoding, encoding_errors="replace",
        keep_default_na=False, skiprows=skip_rows,
        chunksize=CHUNK_ROWS, on_bad_lines="skip",
    )

    # Write CSV header row as the first block
    header_bytes = (",".join(headers) + "\n").encode("utf-8")
    block_writer.write(header_bytes)

    original_rows = 0
    clean_rows    = 0
    small_chunks: list[pd.DataFrame] = []

    for chunk in reader:
        original_rows += len(chunk)
        chunk = _clean_chunk(chunk, converters)

        if is_large:
            buf = io.StringIO()
            chunk.to_csv(buf, index=False, header=False)
            block_writer.write(buf.getvalue().encode("utf-8"))
            clean_rows += len(chunk)
        else:
            small_chunks.append(chunk)

    if not is_large and small_chunks:
        full   = pd.concat(small_chunks, ignore_index=True)
        before = len(full)
        full   = full.drop_duplicates()
        n_dup  = before - len(full)
        if n_dup:
            warns.append(f"Dropped {n_dup} exact-duplicate row(s)")
        full = full.fillna("")
        buf  = io.StringIO()
        full.to_csv(buf, index=False, header=False)
        block_writer.write(buf.getvalue().encode("utf-8"))
        clean_rows = len(full)

    block_writer.commit()

    return {
        "original_rows": original_rows,
        "clean_rows":    clean_rows,
        "cols_renamed":  cols_renamed,
        "encoding":      encoding,
    }


def _clean_chunk(chunk: pd.DataFrame, converters: dict) -> pd.DataFrame:
    """Apply all per-row / per-cell cleaning to a single chunk."""
    chunk = chunk.apply(_clean_str_series).apply(_nullify_series)
    chunk = chunk.dropna(how="all")
    chunk, _ = _drop_garbage_rows(chunk)
    for col, fn in converters.items():
        if col in chunk.columns:
            try:
                chunk[col] = chunk[col].apply(fn)
            except Exception:
                pass
    return chunk.fillna("")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — one temp file download, direct Azure block blob output
# ══════════════════════════════════════════════════════════════════════════════

def _process_excel_to_blob(
    raw_path:     str,
    block_writer: _BlockBlobWriter,
    ext:          str,
    is_large:     bool,
    warns:        list[str],
) -> dict:
    if ext in (".xlsx", ".xlsm"):
        return _process_xlsx_to_blob(raw_path, block_writer, is_large, warns)
    return _process_xls_to_blob(raw_path, block_writer, warns)


def _process_xlsx_to_blob(
    raw_path:     str,
    block_writer: _BlockBlobWriter,
    is_large:     bool,
    warns:        list[str],
) -> dict:
    """
    Stream .xlsx via openpyxl read_only=True, write directly to Azure block blob.
    Peak disk: only the .xlsx download (no second clean file).
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string  # noqa: PLC0415

    wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    ws = wb.active

    hidden_col_idx: set[int] = set()
    try:
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.hidden:
                try:
                    hidden_col_idx.add(column_index_from_string(col_letter))
                except Exception:
                    pass
    except Exception:
        pass

    def _iter_visible_rows(ws_obj) -> Iterator[list[str]]:
        for row in ws_obj.iter_rows():
            row_num = row[0].row if row else None
            if row_num is None:
                continue
            try:
                rd = ws_obj.row_dimensions.get(row_num)
                if rd and rd.hidden:
                    continue
            except Exception:
                pass
            yield [
                str(cell.value) if cell.value is not None else ""
                for j, cell in enumerate(row, start=1)
                if j not in hidden_col_idx
            ]

    # Collect header-scan rows
    head_buf: list[list[str]] = []
    row_iter = _iter_visible_rows(ws)
    for raw_row in row_iter:
        head_buf.append(raw_row)
        if len(head_buf) >= HEADER_SCAN_ROWS:
            break
    wb.close()

    if not head_buf:
        warns.append("Excel file appears empty")
        block_writer.write(b"")
        block_writer.commit()
        return {"original_rows": 0, "clean_rows": 0, "cols_renamed": {}, "encoding": "binary"}

    head_df        = pd.DataFrame(head_buf).astype(str)
    head_df        = head_df.apply(_clean_str_series).apply(_nullify_series)
    header_row_idx = _find_header_row(head_df)
    raw_headers    = [
        _flatten_col_name(v) or f"col_{i}"
        for i, v in enumerate(head_df.iloc[header_row_idx])
    ]
    headers, cols_renamed = _dedup_column_names(raw_headers)
    n_cols        = len(headers)
    data_leftover = head_buf[header_row_idx + 1:]

    # Re-open for the full streaming pass
    wb2 = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    ws2 = wb2.active

    def _iter_data_rows() -> Iterator[list[str]]:
        for r in data_leftover:
            yield r[:n_cols] + [""] * max(0, n_cols - len(r))
        skipped   = 0
        skip_limit = header_row_idx + 1
        for row in ws2.iter_rows():
            row_num = row[0].row if row else None
            if row_num is None:
                continue
            try:
                rd = ws2.row_dimensions.get(row_num)
                if rd and rd.hidden:
                    continue
            except Exception:
                pass
            if skipped < skip_limit:
                skipped += 1
                continue
            vals = [
                str(cell.value) if cell.value is not None else ""
                for j, cell in enumerate(row, start=1)
                if j not in hidden_col_idx
            ]
            yield vals[:n_cols] + [""] * max(0, n_cols - len(vals))

    # Sample for type detection
    sample_rows: list[list[str]] = []
    data_gen = _iter_data_rows()
    for row in data_gen:
        sample_rows.append(row)
        if len(sample_rows) >= TYPE_DETECT_SAMPLE_ROWS:
            break

    sample_df  = pd.DataFrame(sample_rows, columns=headers).astype(str)
    sample_df  = sample_df.apply(_clean_str_series).apply(_nullify_series)
    converters = _build_converters(sample_df, headers, warns)

    # Write header to block blob
    block_writer.write((",".join(headers) + "\n").encode("utf-8"))

    original_rows = len(sample_rows)
    clean_rows    = 0
    small_chunks: list[pd.DataFrame] = []

    def _make_chunk(rows: list[list[str]]) -> pd.DataFrame:
        df = pd.DataFrame(rows, columns=headers).astype(str)
        return _clean_chunk(df, converters)

    sample_clean = _make_chunk(sample_rows)
    if is_large:
        buf = io.StringIO()
        sample_clean.to_csv(buf, index=False, header=False)
        block_writer.write(buf.getvalue().encode("utf-8"))
        clean_rows += len(sample_clean)
    else:
        small_chunks.append(sample_clean)

    batch: list[list[str]] = []
    for row in data_gen:
        original_rows += 1
        batch.append(row)
        if len(batch) >= CHUNK_ROWS:
            chunk = _make_chunk(batch)
            if is_large:
                buf = io.StringIO()
                chunk.to_csv(buf, index=False, header=False)
                block_writer.write(buf.getvalue().encode("utf-8"))
                clean_rows += len(chunk)
            else:
                small_chunks.append(chunk)
            batch = []

    if batch:
        chunk = _make_chunk(batch)
        if is_large:
            buf = io.StringIO()
            chunk.to_csv(buf, index=False, header=False)
            block_writer.write(buf.getvalue().encode("utf-8"))
            clean_rows += len(chunk)
        else:
            small_chunks.append(chunk)

    if not is_large and small_chunks:
        full   = pd.concat(small_chunks, ignore_index=True)
        before = len(full)
        full   = full.drop_duplicates()
        n_dup  = before - len(full)
        if n_dup:
            warns.append(f"Dropped {n_dup} exact-duplicate row(s)")
        full = full.fillna("")
        buf  = io.StringIO()
        full.to_csv(buf, index=False, header=False)
        block_writer.write(buf.getvalue().encode("utf-8"))
        clean_rows = len(full)

    block_writer.commit()
    wb2.close()

    return {
        "original_rows": original_rows,
        "clean_rows":    clean_rows,
        "cols_renamed":  cols_renamed,
        "encoding":      "binary",
    }


def _process_xls_to_blob(
    raw_path:     str,
    block_writer: _BlockBlobWriter,
    warns:        list[str],
) -> dict:
    """
    .xls legacy format via xlrd + pandas.
    xlrd doesn't support streaming but .xls is capped at 65 535 rows.
    Clean output written directly to block blob — no second local file.
    """
    try:
        df = pd.read_excel(raw_path, header=None, dtype=object, engine="xlrd")
    except ImportError:
        warns.append("xlrd not installed; .xls support unavailable")
        block_writer.commit()
        return {"original_rows": 0, "clean_rows": 0, "cols_renamed": {}, "encoding": "binary"}
    except Exception as ex:
        warns.append(f".xls read failed: {ex}")
        block_writer.commit()
        return {"original_rows": 0, "clean_rows": 0, "cols_renamed": {}, "encoding": "binary"}

    original_rows  = len(df)
    df = df.astype(str).apply(_clean_str_series).apply(_nullify_series)

    header_row_idx = _find_header_row(df.iloc[:HEADER_SCAN_ROWS])
    raw_headers    = [
        _flatten_col_name(v) or f"col_{i}"
        for i, v in enumerate(df.iloc[header_row_idx])
    ]
    headers, cols_renamed = _dedup_column_names(raw_headers)

    df = df.iloc[header_row_idx + 1:].copy()
    ncols = len(df.columns)
    df.columns = headers[:ncols] + [f"col_{i}" for i in range(ncols - len(headers))]
    converters = _build_converters(df.iloc[:TYPE_DETECT_SAMPLE_ROWS], headers, warns)
    df = _clean_chunk(df, converters)

    before = len(df)
    df = df.drop_duplicates()
    n_dup = before - len(df)
    if n_dup:
        warns.append(f"Dropped {n_dup} exact-duplicate row(s)")

    buf = io.StringIO()
    df.fillna("").to_csv(buf, index=False)
    block_writer.write(buf.getvalue().encode("utf-8"))
    block_writer.commit()

    return {
        "original_rows": original_rows,
        "clean_rows":    len(df),
        "cols_renamed":  cols_renamed,
        "encoding":      "binary",
    }


# ══════════════════════════════════════════════════════════════════════════════
# Encoding + delimiter detection
# ══════════════════════════════════════════════════════════════════════════════

def _detect_encoding_from_bytes(raw: bytes) -> str:
    """Detect encoding from a bytes object (e.g. probe data already in memory)."""
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if raw[:4] in (b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff"):
        return "utf-32"
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    try:
        from charset_normalizer import from_bytes as _cnb  # noqa: PLC0415
        best = _cnb(raw).best()
        if best:
            return str(best.encoding)
    except ImportError:
        pass
    for enc in ("utf-8", "cp1252", "iso-8859-1", "latin-1"):
        try:
            raw.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            pass
    return "utf-8"


def _detect_encoding(path: str) -> str:
    """Detect encoding from a local file path (used by the file-path test helpers)."""
    with open(path, "rb") as fh:
        return _detect_encoding_from_bytes(fh.read(65536))


def _detect_delimiter_from_str(text: str, ext_hint: str = "") -> str:
    """Detect CSV delimiter from already-decoded text (e.g. probe text in memory)."""
    try:
        return csv.Sniffer().sniff(text[:8192], delimiters=",;\t|").delimiter
    except csv.Error:
        return "\t" if ext_hint.endswith((".tsv", ".tab")) else ","


def _detect_delimiter(path: str, encoding: str) -> str:
    """Detect CSV delimiter from a local file path (used by the file-path test helpers)."""
    try:
        with open(path, encoding=encoding, errors="replace") as fh:
            return _detect_delimiter_from_str(fh.read(8192), path)
    except OSError:
        return ","


# ══════════════════════════════════════════════════════════════════════════════
# File-path based processing (kept for tests + local fallback)
# ══════════════════════════════════════════════════════════════════════════════

def _process_text(
    raw_path:   str,
    clean_path: str,
    is_large:   bool,
    warns:      list[str],
) -> dict:
    """
    File-path based CSV processing (used in unit tests and local fallback).
    For production Azure ingestion, _process_text_stream is used instead.
    """
    encoding  = _detect_encoding(raw_path)
    delimiter = _detect_delimiter(raw_path, encoding)

    head_df = pd.read_csv(
        raw_path, sep=delimiter, header=None, dtype=str,
        encoding=encoding, encoding_errors="replace",
        keep_default_na=False, nrows=HEADER_SCAN_ROWS,
        on_bad_lines="skip",
    )
    head_df = head_df.apply(_clean_str_series).apply(_nullify_series)
    header_row_idx = _find_header_row(head_df)

    raw_headers = [
        _flatten_col_name(v) or f"col_{i}"
        for i, v in enumerate(head_df.iloc[header_row_idx])
    ]
    headers, cols_renamed = _dedup_column_names(raw_headers)
    skip_rows = header_row_idx + 1

    if header_row_idx > 0:
        warns.append(f"Header row found at row {header_row_idx} ({header_row_idx} leading rows skipped)")

    sample_df = pd.read_csv(
        raw_path, sep=delimiter, header=None, dtype=str, names=headers,
        encoding=encoding, encoding_errors="replace",
        keep_default_na=False, skiprows=skip_rows,
        nrows=TYPE_DETECT_SAMPLE_ROWS, on_bad_lines="skip",
    )
    sample_df = sample_df.apply(_clean_str_series).apply(_nullify_series)
    converters = _build_converters(sample_df, headers, warns)

    original_rows = 0
    clean_rows    = 0
    small_chunks: list[pd.DataFrame] = []

    reader = pd.read_csv(
        raw_path, sep=delimiter, header=None, dtype=str, names=headers,
        encoding=encoding, encoding_errors="replace",
        keep_default_na=False, skiprows=skip_rows,
        chunksize=CHUNK_ROWS, on_bad_lines="skip",
    )

    with open(clean_path, "w", encoding="utf-8", newline="") as out_fh:
        writer = csv.writer(out_fh, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)

        for chunk in reader:
            original_rows += len(chunk)
            chunk = _clean_chunk(chunk, converters)

            if is_large:
                for row in chunk.itertuples(index=False, name=None):
                    writer.writerow(row)
                clean_rows += len(chunk)
            else:
                small_chunks.append(chunk)

        if not is_large and small_chunks:
            full = pd.concat(small_chunks, ignore_index=True)
            before = len(full)
            full   = full.drop_duplicates()
            n_dup  = before - len(full)
            if n_dup:
                warns.append(f"Dropped {n_dup} exact-duplicate row(s)")
            full = full.fillna("")
            for row in full.itertuples(index=False, name=None):
                writer.writerow(row)
            clean_rows = len(full)

    return {
        "original_rows": original_rows,
        "clean_rows":    clean_rows,
        "cols_renamed":  cols_renamed,
        "encoding":      encoding,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Cell-level string cleaning
# ══════════════════════════════════════════════════════════════════════════════

def _clean_str(v: object) -> object:
    if not isinstance(v, str):
        return v
    v = _INVISIBLE_RE.sub("", v)
    v = unicodedata.normalize("NFKC", v)
    v = _CTRL_RE.sub("", v)
    v = v.replace("\r\n", " ").replace("\r", " ").replace("\n", " ").replace("\t", " ")
    v = re.sub(r" {2,}", " ", v).strip()
    return v if v else None


def _clean_str_series(s: pd.Series) -> pd.Series:
    return s.apply(_clean_str)


def _nullify(v: object) -> object:
    if v is None:
        return None
    if isinstance(v, float) and np.isnan(v):
        return None
    sv = str(v).strip().lower()
    return None if sv in _NULLSTR else v


def _nullify_series(s: pd.Series) -> pd.Series:
    return s.apply(_nullify)


# ══════════════════════════════════════════════════════════════════════════════
# Header detection (reads only first HEADER_SCAN_ROWS)
# ══════════════════════════════════════════════════════════════════════════════

def _is_numeric_str(v: str) -> bool:
    try:
        float(str(v).strip().replace(",", "").replace("_", ""))
        return True
    except (ValueError, TypeError):
        return False


def _find_header_row(df: pd.DataFrame) -> int:
    max_scan  = min(HEADER_SCAN_ROWS, len(df))
    best_row, best_score = 0, -1.0

    for i in range(max_scan):
        row      = df.iloc[i]
        non_null = [v for v in row if v is not None]
        if not non_null:
            continue
        str_cnt   = sum(1 for v in non_null if isinstance(v, str) and not _is_numeric_str(v))
        num_cnt   = sum(1 for v in non_null if isinstance(v, str) and _is_numeric_str(v))
        coverage  = len(non_null) / max(len(row), 1)
        str_ratio = str_cnt / max(len(non_null), 1)
        num_ratio = num_cnt / max(len(non_null), 1)
        avg_len   = sum(len(str(v)) for v in non_null) / max(len(non_null), 1)
        len_pen   = max(0.0, (avg_len - 60) / 150)
        score = coverage * str_ratio - num_ratio * 0.5 - len_pen

        if score > best_score + 0.05:
            best_score = score
            best_row   = i
        if i <= 3 and score > 0.70:
            break

    return best_row


# ══════════════════════════════════════════════════════════════════════════════
# Structural helpers
# ══════════════════════════════════════════════════════════════════════════════

def _flatten_col_name(v: object) -> str:
    if isinstance(v, tuple):
        parts = [str(p).strip() for p in v
                 if p is not None and not str(p).lower().startswith("unnamed")]
        return "_".join(parts) if parts else ""
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\s*Unnamed:\s*\d+(_level_\d+)?", "", s, flags=re.IGNORECASE).strip()
    return s


def _dedup_column_names(names: list[str]) -> tuple[list[str], dict]:
    renamed: dict[str, str] = {}
    seen:    dict[str, int] = {}
    result:  list[str]      = []
    for col in names:
        clean = col.strip() or "col"
        if clean in seen:
            seen[clean] += 1
            new = f"{clean}_{seen[clean]}"
            renamed[col] = new
            result.append(new)
        else:
            seen[clean] = 0
            result.append(clean)
    return result, renamed


def _drop_garbage_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    keep = []
    for _, row in df.iterrows():
        non_null = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if not non_null:
            keep.append(False)
            continue
        first = non_null[0]
        if all(_SEP_ROW_RE.match(c) for c in non_null):
            keep.append(False)
        elif _GARBAGE_ROW_RE.match(first):
            keep.append(False)
        else:
            keep.append(True)
    mask   = pd.Series(keep, index=df.index)
    n_drop = int((~mask).sum())
    return df[mask].copy(), n_drop


# ══════════════════════════════════════════════════════════════════════════════
# Type detection + converter builder (uses only the sample; applied to all chunks)
# ══════════════════════════════════════════════════════════════════════════════

ConverterFn = Callable[[object], object]


def _build_converters(
    sample: pd.DataFrame, headers: list[str], warns: list[str],
) -> dict[str, ConverterFn]:
    """
    Inspect sample rows ONCE and return {col: converter_fn}.
    The returned functions are applied cell-by-cell to every chunk —
    they do not accumulate state across chunks.
    """
    converters: dict[str, ConverterFn] = {}
    for col in headers:
        if col not in sample.columns:
            continue
        series    = sample[col].dropna()
        col_lower = col.lower()
        if series.empty:
            continue

        unique_lower = {str(v).strip().lower() for v in series}

        # ── Bool ──────────────────────────────────────────────────────────────
        if unique_lower and unique_lower.issubset(_BOOL_ALL):
            converters[col] = _make_bool_converter()
            continue

        # ── Date ──────────────────────────────────────────────────────────────
        is_date_hint = any(h in col_lower for h in _DATE_NAME_CLUES)
        threshold    = DATE_CONVERT_THRESHOLD if is_date_hint else 0.80
        ratio        = _date_parse_ratio(series)
        if ratio >= threshold:
            n_failed = int(len(series) * (1 - ratio))
            if n_failed:
                warns.append(
                    f"Column '{col}': ~{n_failed} date value(s) in sample could not be parsed"
                )
            converters[col] = _make_date_converter()
            continue

        # ── Numeric ───────────────────────────────────────────────────────────
        num_thresh = (NUM_HINT_THRESHOLD
                      if any(h in col_lower for h in _NUM_NAME_CLUES)
                      else NUM_CONVERT_THRESHOLD)
        if _numeric_parse_ratio(series) >= num_thresh:
            converters[col] = _make_numeric_converter()
            continue

    return converters


# ── Bool ──────────────────────────────────────────────────────────────────────

def _make_bool_converter() -> ConverterFn:
    def _fn(v: object) -> object:
        if v is None:
            return None
        sv = str(v).strip().lower()
        if sv in _BOOL_TRUE:  return "True"
        if sv in _BOOL_FALSE: return "False"
        return None
    return _fn


# ── Date ──────────────────────────────────────────────────────────────────────

def _excel_serial_to_iso(n: float) -> str | None:
    try:
        ni = int(n)
        if not (_EXCEL_SERIAL_MIN <= ni <= _EXCEL_SERIAL_MAX):
            return None
        dt = _EXCEL_EPOCH + timedelta(days=ni)
        return dt.strftime("%Y-%m-%d") if 1900 <= dt.year <= 2100 else None
    except (ValueError, OverflowError, OSError):
        return None


def _parse_one_date(v: object) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d") if 1900 <= v.year <= 2100 else None
    if isinstance(v, _date_type):
        return v.isoformat() if 1900 <= v.year <= 2100 else None
    sv = str(v).strip()
    if not sv or sv.lower() in _NULLSTR:
        return None
    try:
        f = float(sv.replace(",", ""))
        if f == int(f):
            result = _excel_serial_to_iso(f)
            if result:
                return result
    except (ValueError, TypeError):
        pass
    try:
        from dateutil import parser as _dp  # noqa: PLC0415
        parsed = _dp.parse(sv, default=datetime(1900, 1, 1), dayfirst=False)
        return parsed.strftime("%Y-%m-%d") if 1900 <= parsed.year <= 2100 else None
    except Exception:
        pass
    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d",
        "%d-%m-%Y", "%m-%d-%Y", "%d.%m.%Y", "%m.%d.%Y",
        "%Y%m%d", "%d %b %Y", "%b %d %Y", "%d %B %Y", "%B %d %Y",
        "%b-%d-%Y", "%d-%b-%Y", "%b %Y", "%B %Y",
    ):
        try:
            parsed = datetime.strptime(sv, fmt)
            return parsed.strftime("%Y-%m-%d") if 1900 <= parsed.year <= 2100 else None
        except ValueError:
            pass
    return None


def _date_parse_ratio(series: pd.Series) -> float:
    return series.apply(_parse_one_date).notna().sum() / max(len(series), 1)


def _make_date_converter() -> ConverterFn:
    def _fn(v: object) -> object:
        return _parse_one_date(v)
    return _fn


# ── Numeric ───────────────────────────────────────────────────────────────────

def _strip_numeric_noise(raw: str) -> str | None:
    v = _INVISIBLE_RE.sub("", raw.strip())
    v = unicodedata.normalize("NFKC", v)
    m = _PERCENT_RE.match(v)
    if m:
        try:
            return str(round(float(m.group(1)) / 100, 12))
        except ValueError:
            pass
    v = _CURRENCY_RE.sub("", v).strip()
    v = _SPACE_THOU_RE.sub(r"\1\2", v)
    if _COMMA_THOU_RE.match(v):
        v = v.replace(",", "")
    return v.strip() or None


def _numeric_parse_ratio(series: pd.Series) -> float:
    def _parseable(v: object) -> bool:
        if v is None:
            return False
        cleaned = _strip_numeric_noise(str(v))
        if not cleaned:
            return False
        try:
            float(cleaned)
            return True
        except (ValueError, TypeError):
            return False
    return series.apply(_parseable).sum() / max(len(series), 1)


def _make_numeric_converter() -> ConverterFn:
    def _fn(v: object) -> object:
        if v is None:
            return None
        sv = str(v).strip()
        if not sv:
            return None
        cleaned = _strip_numeric_noise(sv)
        if not cleaned:
            return v
        try:
            f = float(cleaned)
            if np.isnan(f) or np.isinf(f):
                return None
            if f == int(f) and "." not in cleaned:
                return str(int(f))
            return str(f)
        except (ValueError, OverflowError):
            return v
    return _fn
